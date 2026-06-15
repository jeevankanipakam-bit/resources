import boto3
from botocore.exceptions import ClientError
from openpyxl import Workbook

wb = Workbook()
wb.remove(wb.active)


def get_all_regions():
    """Dynamically gets all enabled regions for your AWS account."""
    try:
        # Using us-east-1 just to fetch the region list
        ec2 = boto3.client("ec2", region_name="us-east-1")
        regions = [
            reg["RegionName"]
            for reg in ec2.describe_regions()["Regions"]
            if reg["OptInStatus"] != "not-opted-in"
        ]
        return regions
    except Exception as e:
        print(f"Error fetching regions, falling back to basic list: {e}")
        return ["us-east-1", "us-west-2", "eu-west-1", "ap-south-1", "ap-south-2"]


def create_sheet(sheet_name, headers, rows):
    ws = wb.create_sheet(title=sheet_name[:31])
    ws.append(headers)

    for row in rows:
        ws.append(row)

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 5, 50)


def safe_call(func, region=None):
    try:
        if region:
            return func(region)
        return func()
    except Exception as e:
        region_str = f" in {region}" if region else ""
        print(f"Error executing {func.__name__}{region_str}: {e}")
        return []


# ==============================================================================
# REGIONAL SERVICES (Requires a region parameter)
# ==============================================================================


# EC2
def get_ec2_instances(region):
    ec2 = boto3.client("ec2", region_name=region)
    rows = []
    paginator = ec2.get_paginator("describe_instances")

    for page in paginator.paginate():
        for reservation in page["Reservations"]:
            for instance in reservation["Instances"]:
                name = ""
                for tag in instance.get("Tags", []):
                    if tag["Key"] == "Name":
                        name = tag["Value"]
                        break

                rows.append(
                    [
                        region,
                        instance.get("InstanceId"),
                        name,
                        instance.get("InstanceType"),
                        instance.get("State", {}).get("Name"),
                        instance.get("VpcId"),
                        instance.get("SubnetId"),
                        instance.get("PrivateIpAddress"),
                        instance.get("PublicIpAddress"),
                        instance.get("KeyName"),
                        instance.get("ImageId"),
                        instance.get("Placement", {}).get("AvailabilityZone"),
                    ]
                )
    return rows


# EBS Volumes
def get_ebs_volumes(region):
    ec2 = boto3.client("ec2", region_name=region)
    rows = []
    for vol in ec2.describe_volumes()["Volumes"]:
        rows.append([region, vol["VolumeId"], vol["Size"], vol["State"], vol.get("VolumeType")])
    return rows


# Snapshots
def get_snapshots(region):
    ec2 = boto3.client("ec2", region_name=region)
    rows = []
    paginator = ec2.get_paginator("describe_snapshots")

    for page in paginator.paginate(OwnerIds=["self"]):
        for snap in page["Snapshots"]:
            rows.append(
                [
                    region,
                    snap["SnapshotId"],
                    snap["VolumeId"],
                    snap["State"],
                    snap["StartTime"].strftime("%Y-%m-%d"),
                ]
            )
    return rows


# AMIs
def get_amis(region):
    ec2 = boto3.client("ec2", region_name=region)
    rows = []
    images = ec2.describe_images(Owners=["self"])["Images"]

    for img in images:
        rows.append([region, img["ImageId"], img.get("Name"), img.get("State")])
    return rows


# Launch Templates
def get_launch_templates(region):
    ec2 = boto3.client("ec2", region_name=region)
    rows = []
    templates = ec2.describe_launch_templates().get("LaunchTemplates", [])

    for lt in templates:
        rows.append([region, lt["LaunchTemplateId"], lt["LaunchTemplateName"]])
    return rows


# Elastic IPs
def get_elastic_ips(region):
    ec2 = boto3.client("ec2", region_name=region)
    rows = []
    for ip in ec2.describe_addresses()["Addresses"]:
        rows.append([region, ip.get("PublicIp"), ip.get("AllocationId"), ip.get("AssociationId")])
    return rows


# Key Pairs
def get_keypairs(region):
    ec2 = boto3.client("ec2", region_name=region)
    rows = []
    for kp in ec2.describe_key_pairs()["KeyPairs"]:
        rows.append([region, kp["KeyName"], kp["KeyType"]])
    return rows


# Load Balancers
def get_load_balancers(region):
    elb = boto3.client("elbv2", region_name=region)
    rows = []
    for lb in elb.describe_load_balancers()["LoadBalancers"]:
        rows.append([region, lb["LoadBalancerName"], lb["Type"], lb["State"]["Code"]])
    return rows


# Target Groups
def get_target_groups(region):
    elb = boto3.client("elbv2", region_name=region)
    rows = []
    for tg in elb.describe_target_groups()["TargetGroups"]:
        rows.append([region, tg["TargetGroupName"], tg["Protocol"], tg["Port"]])
    return rows


# Auto Scaling Groups
def get_asgs(region):
    asg = boto3.client("autoscaling", region_name=region)
    rows = []
    paginator = asg.get_paginator("describe_auto_scaling_groups")

    for page in paginator.paginate():
        for group in page["AutoScalingGroups"]:
            rows.append(
                [
                    region,
                    group["AutoScalingGroupName"],
                    group["DesiredCapacity"],
                    group["MinSize"],
                    group["MaxSize"],
                ]
            )
    return rows


# Security Groups
def get_security_groups(region):
    ec2 = boto3.client("ec2", region_name=region)
    return [
        [region, sg["GroupId"], sg["GroupName"], sg["VpcId"]]
        for sg in ec2.describe_security_groups()["SecurityGroups"]
    ]


# VPCs
def get_vpcs(region):
    ec2 = boto3.client("ec2", region_name=region)
    return [
        [region, v["VpcId"], v["CidrBlock"], v["State"]] for v in ec2.describe_vpcs()["Vpcs"]
    ]


# Subnets
def get_subnets(region):
    ec2 = boto3.client("ec2", region_name=region)
    rows = []
    for subnet in ec2.describe_subnets()["Subnets"]:
        rows.append(
            [
                region,
                subnet["SubnetId"],
                subnet["VpcId"],
                subnet["AvailabilityZone"],
                subnet["MapPublicIpOnLaunch"],
            ]
        )
    return rows


# EFS
def get_efs(region):
    efs = boto3.client("efs", region_name=region)
    return [
        [region, fs["FileSystemId"], fs["Name"] if "Name" in fs else "", fs["LifeCycleState"]]
        for fs in efs.describe_file_systems()["FileSystems"]
    ]


# Secrets Manager
# def get_secrets(region):
#     sm = boto3.client("secretsmanager", region_name=region)
#     return [[region, s["Name"], s["ARN"]] for s in sm.list_secrets()["SecretList"]]


# RDS & Aurora
def get_rds(region):
    rds = boto3.client("rds", region_name=region)
    return [
        [region, db["DBInstanceIdentifier"], db["Engine"], db["DBInstanceStatus"]]
        for db in rds.describe_db_instances()["DBInstances"]
    ]


# DynamoDB
def get_dynamodb(region):
    ddb = boto3.client("dynamodb", region_name=region)
    return [[region, table] for table in ddb.list_tables()["TableNames"]]


# ElastiCache
# def get_elasticache(region):
#     ec = boto3.client("elasticache", region_name=region)
#     return [[region, c["CacheClusterId"], c["Engine"]] for c in ec.describe_cache_clusters()["CacheClusters"]]


# ECR
def get_ecr(region):
    ecr = boto3.client("ecr", region_name=region)
    return [
        [region, repo["repositoryName"], repo["repositoryUri"]]
        for repo in ecr.describe_repositories()["repositories"]
    ]


# ECS
# def get_ecs(region):
#     ecs = boto3.client("ecs", region_name=region)
#     return [[region, cluster] for cluster in ecs.list_clusters()["clusterArns"]]


# EKS
def get_eks(region):
    eks = boto3.client("eks", region_name=region)
    return [[region, cluster] for cluster in eks.list_clusters()["clusters"]]


# Redshift
def get_redshift(region):
    rs = boto3.client("redshift", region_name=region)
    return [
        [region, c["ClusterIdentifier"], c["NodeType"]] for c in rs.describe_clusters()["Clusters"]
    ]


# Lambda
def get_lambda(region):
    lm = boto3.client("lambda", region_name=region)
    return [
        [region, f["FunctionName"], f["Runtime"]] for f in lm.list_functions()["Functions"]
    ]


# Step Functions
def get_stepfunctions(region):
    sf = boto3.client("stepfunctions", region_name=region)
    return [
        [region, sm["name"], sm["stateMachineArn"]]
        for sm in sf.list_state_machines()["stateMachines"]
    ]


# CloudFormation
def get_cloudformation(region):
    cfn = boto3.client("cloudformation", region_name=region)
    return [
        [region, s["StackName"], s["StackStatus"]] for s in cfn.describe_stacks()["Stacks"]
    ]


# EventBridge
def get_eventbridge(region):
    events = boto3.client("events", region_name=region)
    rows = []
    paginator = events.get_paginator("list_rules")

    for page in paginator.paginate():
        for rule in page["Rules"]:
            rows.append([region, rule.get("Name"), rule.get("State"), rule.get("EventBusName")])
    return rows


# CloudWatch Alarms
def get_cloudwatch_alarms(region):
    cw = boto3.client("cloudwatch", region_name=region)
    rows = []
    paginator = cw.get_paginator("describe_alarms")

    for page in paginator.paginate():
        for alarm in page.get("MetricAlarms", []):
            rows.append(
                [region, alarm.get("AlarmName"), alarm.get("StateValue"), alarm.get("MetricName")]
            )
    return rows


# WAF Regional
# def get_waf(region):
#     rows = []
#     try:
#         waf = boto3.client("wafv2", region_name=region)
#         response = waf.list_web_acls(Scope="REGIONAL", Limit=100)
#         for acl in response.get("WebACLs", []):
#             rows.append([region, acl["Name"], acl["Id"], "REGIONAL"])
#     except Exception as e:
#         print(f"WAF Error: {e}")
#     return rows


# Network ACLs
# def get_nacls(region):
#     ec2 = boto3.client("ec2", region_name=region)
#     rows = []
#     for nacl in ec2.describe_network_acls()["NetworkAcls"]:
#         rows.append([region, nacl["NetworkAclId"], nacl["VpcId"], nacl["IsDefault"]])
#     return rows


# CloudTrail
# def get_cloudtrail(region):
#     ct = boto3.client("cloudtrail", region_name=region)
#     rows = []
#     for trail in ct.describe_trails()["trailList"]:
#         rows.append([region, trail.get("Name"), trail.get("HomeRegion"), trail.get("S3BucketName")])
#     return rows


# KMS
# def get_kms(region):
#     kms = boto3.client("kms", region_name=region)
#     rows = []
#     paginator = kms.get_paginator("list_keys")
#     for page in paginator.paginate():
#         for key in page["Keys"]:
#             try:
#                 metadata = kms.describe_key(KeyId=key["KeyId"])["KeyMetadata"]
#                 rows.append([region, metadata["KeyId"], metadata["Arn"], metadata["KeyManager"], metadata["KeyState"]])
#             except Exception:
#                 pass
#     return rows


# ==============================================================================
# GLOBAL SERVICES (Does NOT change data per region loop)
# ==============================================================================


# S3
def get_s3():
    s3 = boto3.client("s3")
    return [[bucket["Name"]] for bucket in s3.list_buckets()["Buckets"]]


# IAM Roles
def get_iam_roles():
    iam = boto3.client("iam")
    return [[r["RoleName"]] for r in iam.list_roles()["Roles"]]


# IAM Policies
def get_iam_policies():
    iam = boto3.client("iam")
    return [[p["PolicyName"]] for p in iam.list_policies(Scope="Local")["Policies"]]


# Route53
def get_route53():
    r53 = boto3.client("route53")
    return [[z["Name"], z["Id"]] for z in r53.list_hosted_zones()["HostedZones"]]


# CloudFront
def get_cloudfront():
    cf = boto3.client("cloudfront")
    rows = []
    response = cf.list_distributions()

    for dist in response.get("DistributionList", {}).get("Items", []):
        rows.append(
            [
                dist.get("Id"),
                dist.get("DomainName"),
                dist.get("Status"),
                ",".join(dist.get("Aliases", {}).get("Items", [])),
            ]
        )
    return rows


# WAF CloudFront
# def get_waf_cloudfront():
#     rows = []
#     try:
#         waf = boto3.client("wafv2", region_name="us-east-1")
#         response = waf.list_web_acls(Scope="CLOUDFRONT", Limit=100)
#         for acl in response.get("WebACLs", []):
#             rows.append([acl["Name"], acl["Id"], "CLOUDFRONT"])
#     except Exception as e:
#         print(f"WAF CloudFront Error: {e}")
#     return rows


# Shield
# def get_shield():
#     rows = []
#     try:
#         shield = boto3.client("shield")
#         response = shield.list_protections()
#         for protection in response.get("Protections", []):
#             rows.append([protection.get("Name"), protection.get("Id"), protection.get("ResourceArn")])
#     except Exception as e:
#         print(f"Shield Error: {e}")
#     return rows


# ==============================================================================
# EXECUTION ROUTINE
# ==============================================================================

# Differentiate execution maps
regional_resources = [
    (
        "EC2",
        [
            "Region",
            "InstanceId",
            "Name",
            "InstanceType",
            "State",
            "VpcId",
            "SubnetId",
            "PrivateIP",
            "PublicIP",
            "KeyPair",
            "AMI",
            "AvailabilityZone",
        ],
        get_ec2_instances,
    ),
    ("EBS", ["Region", "VolumeId", "Size", "State", "Type"], get_ebs_volumes),
    ("Snapshots", ["Region", "SnapshotId", "VolumeId", "State", "Date"], get_snapshots),
    ("AMI", ["Region", "ImageId", "Name", "State"], get_amis),
    ("LaunchTemplates", ["Region", "Id", "Name"], get_launch_templates),
    ("ElasticIPs", ["Region", "PublicIP", "AllocationId", "AssociationId"], get_elastic_ips),
    ("KeyPairs", ["Region", "Name", "Type"], get_keypairs),
    ("LoadBalancers", ["Region", "Name", "Type", "State"], get_load_balancers),
    ("TargetGroups", ["Region", "Name", "Protocol", "Port"], get_target_groups),
    ("AutoScaling", ["Region", "Name", "Desired", "Min", "Max"], get_asgs),
    ("SecurityGroups", ["Region", "GroupId", "Name", "VpcId"], get_security_groups),
    ("VPCs", ["Region", "VpcId", "CIDR", "State"], get_vpcs),
    ("Subnets", ["Region", "SubnetId", "VpcId", "AZ", "Public"], get_subnets),
    ("EFS", ["Region", "FileSystemId", "Name", "State"], get_efs),
    # ("SecretsManager", ["Region", "Name", "ARN"], get_secrets),
    ("RDS", ["Region", "Identifier", "Engine", "Status"], get_rds),
    ("DynamoDB", ["Region", "TableName"], get_dynamodb),
    # ("ElastiCache", ["Region", "Cluster", "Engine"], get_elasticache),
    ("ECR", ["Region", "Repository", "URI"], get_ecr),
    # ("ECS", ["Region", "ClusterArn"], get_ecs),
    ("EKS", ["Region", "ClusterName"], get_eks),
    ("Redshift", ["Region", "Cluster", "NodeType"], get_redshift),
    ("Lambda", ["Region", "FunctionName", "Runtime"], get_lambda),
    ("StepFunctions", ["Region", "Name", "ARN"], get_stepfunctions),
    ("CloudFormation", ["Region", "StackName", "Status"], get_cloudformation),
    ("EventBridge", ["Region", "RuleName", "State", "EventBus"], get_eventbridge),
    ("CloudWatch", ["Region", "AlarmName", "State", "Metric"], get_cloudwatch_alarms),
    # ("WAF", ["Region", "WebACLName", "WebACLId", "Scope"], get_waf),
    # ("NACL", ["Region", "NetworkAclId", "VpcId", "IsDefault"], get_nacls),
    # ("CloudTrail", ["Region", "TrailName", "HomeRegion", "S3Bucket"], get_cloudtrail),
    # ("KMS", ["Region", "KeyId", "ARN", "KeyManager", "KeyState"], get_kms),
]

global_resources = [
    ("S3", ["Bucket"], get_s3),
    ("IAMRoles", ["RoleName"], get_iam_roles),
    ("IAMPolicies", ["PolicyName"], get_iam_policies),
    # ("Route53", ["ZoneName", "ZoneId"], get_route53),
    ("CloudFront", ["DistributionId", "DomainName", "Status", "Aliases"], get_cloudfront),
    # ("WAFCloudFront", ["WebACLName", "WebACLId", "Scope"], get_waf_cloudfront),
    # ("Shield", ["ProtectionName", "ProtectionId", "ResourceArn"], get_shield),
]

# Get list of all target regions
active_regions = get_all_regions()
print(f"Found active regions: {', '.join(active_regions)}\n")

# Process Regional Resources (aggregated into their respective tabs across all regions)
for name, headers, func in regional_resources:
    all_region_rows = []
    print(f"Collecting regional dataset for {name}...")
    for region in active_regions:
        rows = safe_call(func, region=region)
        all_region_rows.extend(rows)
    create_sheet(name, headers, all_region_rows)

# Process Global Resources (run exactly once)
for name, headers, func in global_resources:
    print(f"Collecting global dataset for {name}...")
    rows = safe_call(func)
    create_sheet(name, headers, rows)

output_file = "AWS_Global_Inventory.xlsx"
wb.save(output_file)
print(f"\nAll-region inventory exported successfully: {output_file}")